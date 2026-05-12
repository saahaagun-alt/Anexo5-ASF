"""
Anexo 5 ASF - Web App
======================
Permite generar el Anexo 5 ASF mensual subiendo:
  - Plantilla Anexo 5 (.xlsx)
  - Auxiliar de Cuentas del mes (.xlsx)
  - CFDIs del mes (.xlsx) - opcional
  - Mes
"""

import os
import io
import tempfile
import re
import shutil
import unicodedata
from flask import Flask, request, render_template, send_file, flash, redirect, url_for
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024  # 25 MB
app.secret_key = 'anexo5-municipios-asf'

# ─── CONFIG (copiado de llenar_anexo5.py) ───
MES_COL = {
    'enero': 5, 'febrero': 6, 'marzo': 7, 'abril': 8,
    'mayo': 9, 'junio': 10, 'julio': 11, 'agosto': 12,
    'septiembre': 13, 'octubre': 14, 'noviembre': 15, 'diciembre': 16,
}

ROW_MAP = {
    '113': 9, '122': 11, '132': 25, '133': 25, '152': 30,
    '211': 38, '212': 38, '216': 38, '217': 38, '246': 38, '249': 38,
    '221': 39, '242': 39, '251': 39, '252': 39,
    '253': 40, '261': 41,
    '272': 42, '273': 42, '282': 42, '291': 42, '292': 42,
    '294': 42, '296': 42, '298': 42,
    '311': 45, '313': 45, '314': 45, '315': 45, '317': 45,
    '322': 47, '325': 47, '331': 48, '332': 48, '334': 48, '344': 48, '347': 48,
    '351': 49, '353': 49, '355': 49, '357': 49,
    '375': 50, '382': 52, '392': 52,
    '416': 55, '441': 56,
    '511': 75, '515': 75, '562': 75, '613': 78,
    '911': 81, '921': 82,
}

CAP_NAMES = {
    1000: '1000 SERVICIOS PERSONALES',
    2000: '2000 MATERIALES Y SUMINISTROS',
    3000: '3000 SERVICIOS GENERALES',
    4000: '4000 TRANSFERENCIAS, ASIGNACIONES, SUBSIDIOS Y OTRAS AYUDAS',
    5000: '5000 BIENES MUEBLES, INMUEBLES E INTANGIBLES',
    6000: '6000 INVERSIÓN PÚBLICA',
    9000: '9000 DEUDA PÚBLICA',
}

CAP_SHEET_MAP = {
    1000: ('Detalle de pólizas Cap. 1000', False),
    2000: ('Detalle de pólizas Cap. 2000', True),
    3000: ('Detalle de pólizas Cap. 3000', True),
    4000: ('Detalle de pólizas Cap. 4000', False),
    5000: ('Detalle de pólizas Cap. 5000', True),
    6000: ('Detalle de pólizas Cap. 6000', True),
    9000: ('Detalle de pólizas Cap. 9000', False),
}

NOMBRE_FONDO = {
    '1502': ('FONDO 1502 - FORTASEG (SEGURIDAD PÚBLICA)', 'FORTASEG-1502'),
    '1503': ('FONDO 1503 - PARTICIPACIONES MUNICIPALES (FORTAMUN)', 'FORTAMUN-1503'),
}


def normalize_name(s):
    if not s:
        return ''
    s = str(s).strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    s = s.replace('Ñ', 'N')
    s = re.sub(r'[.,;:]', '', s)
    return re.sub(r'\s+', ' ', s)


def build_rfc_map(cfdi_path):
    if not cfdi_path or not os.path.exists(cfdi_path):
        return {}
    df = pd.read_excel(cfdi_path, sheet_name=0, header=0)
    cols_lower = {c.lower(): c for c in df.columns}
    name_col, rfc_col = None, None
    for k, v in cols_lower.items():
        if 'nombre' in k and 'emisor' in k:
            name_col = v
        if 'rfc' in k and 'emisor' in k:
            rfc_col = v
    if not name_col or not rfc_col:
        return {}
    rfc_map = {}
    for _, row in df.iterrows():
        nombre = row[name_col]
        rfc = row[rfc_col]
        if pd.notna(nombre) and pd.notna(rfc):
            key = normalize_name(nombre)
            if key:
                rfc_map[key] = str(rfc).strip()
    return rfc_map


def match_rfc(beneficiario, rfc_map):
    if not beneficiario:
        return ''
    key = normalize_name(beneficiario)
    if key in rfc_map:
        return rfc_map[key]
    for nombre_norm, rfc in rfc_map.items():
        if (key and nombre_norm and (key in nombre_norm or nombre_norm in key)
                and abs(len(key) - len(nombre_norm)) < 15):
            return rfc
    return ''


def extract_descriptions(file_path):
    df = pd.read_excel(file_path, sheet_name=0, header=None)
    desc_map = {}
    for i in range(len(df)):
        val0 = str(df.iloc[i, 0]) if pd.notna(df.iloc[i, 0]) else ''
        if re.match(r'^8270-\d+-\d+-\d+-\d+$', val0):
            seg5 = val0.split('-')[-1]
            if i + 1 < len(df):
                desc = df.iloc[i + 1, 22]
                if pd.notna(desc):
                    desc_map[seg5] = str(desc).strip()
    return desc_map


def detect_fondo(file_path):
    df = pd.read_excel(file_path, sheet_name=0, header=None)
    for i in range(min(50, len(df))):
        val = str(df.iloc[i, 0]) if pd.notna(df.iloc[i, 0]) else ''
        m = re.match(r'^8270-(\d{4})', val)
        if m:
            return m.group(1)
    return None


def extract_records(file_path, rfc_map):
    df = pd.read_excel(file_path, sheet_name=0, header=None)
    valid_pol = re.compile(r'^[EPDCT]\d{3,}')
    records = []
    current_cuenta = None
    i = 0
    while i < len(df):
        val0 = str(df.iloc[i, 0]) if pd.notna(df.iloc[i, 0]) else ''
        val4 = str(df.iloc[i, 4]) if pd.notna(df.iloc[i, 4]) else ''
        val6 = str(df.iloc[i, 6]) if pd.notna(df.iloc[i, 6]) else ''
        val24 = str(df.iloc[i, 24]) if pd.notna(df.iloc[i, 24]) else ''
        val33 = df.iloc[i, 33] if pd.notna(df.iloc[i, 33]) else None
        if re.match(r'^8270-', val0):
            current_cuenta = val0
        if valid_pol.match(val4) and current_cuenta and val33 is not None:
            benef = ''
            if i + 1 < len(df) and pd.notna(df.iloc[i + 1, 12]):
                benef = str(df.iloc[i + 1, 12]).strip()
            parts = current_cuenta.split('-')
            partida = parts[4][:3] if len(parts) >= 5 else ''
            seg5 = parts[4] if len(parts) >= 5 else ''
            concepto = val24
            if '(' in concepto:
                concepto = concepto[:concepto.index('(')].strip()
            records.append({
                'cuenta': current_cuenta, 'seg5': seg5, 'partida': partida,
                'cap': (int(partida[0]) * 1000) if partida else 0,
                'poliza': val4, 'fecha': val6, 'beneficiario': benef,
                'rfc': match_rfc(benef, rfc_map),
                'concepto': concepto[:80], 'cargo': float(val33),
            })
        i += 1
    return records


# Estilos
THIN = Side(border_style='thin', color='000000')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
DATA_FONT = Font(name='Arial', size=9)
TOTAL_FONT = Font(name='Arial', size=10, bold=True)
TOTAL_FILL = PatternFill('solid', start_color='FFE699')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_MONEY = Alignment(horizontal='right', vertical='center')
MONEY_FMT = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
DATE_FMT = 'dd/mm/yyyy'


def safe_write(ws, row, col, value):
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = mr.bounds
        if min_row <= row <= max_row and min_col <= col <= max_col:
            ws.unmerge_cells(str(mr))
            break
    ws.cell(row=row, column=col, value=value)


def clear_data_area(ws, start_row=17):
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = mr.bounds
        if min_row >= start_row:
            ws.unmerge_cells(str(mr))
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = 'General'


def fill_detalle(ws, cap_recs, has_proveedor, descs, fondo_label, mes_label):
    clear_data_area(ws, start_row=17)
    safe_write(ws, 7, 2, '2025')
    safe_write(ws, 8, 2, '2025')
    safe_write(ws, 9, 2, 'H. AYUNTAMIENTO')
    safe_write(ws, 10, 2, 'TESORERÍA MUNICIPAL')
    safe_write(ws, 11, 2, f'{fondo_label} - {mes_label.upper()} 2025')
    n_cols = 10 if has_proveedor else 8
    importe_col = n_cols
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=15, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[15].height = 35
    start = 17
    for i, r in enumerate(cap_recs):
        row = start + i
        ws.cell(row=row, column=1, value=CAP_NAMES.get(r['cap'], str(r['cap'])))
        ws.cell(row=row, column=2, value=r['seg5'])
        ws.cell(row=row, column=3, value=descs.get(r['seg5'], r['partida']))
        ws.cell(row=row, column=4, value=r['concepto'])
        if has_proveedor:
            ws.cell(row=row, column=5, value=r['beneficiario'])
            ws.cell(row=row, column=6, value=r['rfc'])
            ws.cell(row=row, column=7, value=r['poliza'])
            ws.cell(row=row, column=8, value=r['fecha'])
            ws.cell(row=row, column=9, value=r['fecha'])
            ws.cell(row=row, column=10, value=r['cargo'])
        else:
            ws.cell(row=row, column=5, value=r['poliza'])
            ws.cell(row=row, column=6, value=r['fecha'])
            ws.cell(row=row, column=7, value=r['fecha'])
            ws.cell(row=row, column=8, value=r['cargo'])
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = DATA_FONT
            cell.border = BORDER_ALL
            if c == importe_col:
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT_MONEY
            elif (has_proveedor and c in (8, 9)) or (not has_proveedor and c in (6, 7)):
                cell.number_format = DATE_FMT
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT_WRAP
        ws.row_dimensions[row].height = 30
    total_row = start + len(cap_recs)
    col_letter = get_column_letter(importe_col)
    ws.cell(row=total_row, column=1, value='TOTAL')
    ws.cell(row=total_row, column=importe_col,
            value=f'=SUM({col_letter}17:{col_letter}{total_row-1})')
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = BORDER_ALL
        if c == importe_col:
            cell.number_format = MONEY_FMT
            cell.alignment = RIGHT_MONEY
        else:
            cell.alignment = CENTER
    widths = {1: 28, 2: 12, 3: 32, 4: 45}
    if has_proveedor:
        widths.update({5: 35, 6: 16, 7: 12, 8: 13, 9: 13, 10: 16})
    else:
        widths.update({5: 12, 6: 13, 7: 13, 8: 16})
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def fill_egresos(ws, records, fondo_label, fondo_short, mes_col):
    ws['A2'] = f'CUENTA PÚBLICA 2025 - {fondo_short}'
    ws['A3'] = f'INTEGRACIÓN DE LOS EGRESOS DEL {fondo_label} 2025'
    by_partida = {}
    for r in records:
        by_partida[r['partida']] = by_partida.get(r['partida'], 0.0) + r['cargo']
    row_totals = {}
    not_mapped = {}
    for p, total in by_partida.items():
        if p in ROW_MAP:
            row = ROW_MAP[p]
            row_totals[row] = row_totals.get(row, 0.0) + total
        else:
            not_mapped[p] = total
    for row, total in row_totals.items():
        cell = ws.cell(row=row, column=mes_col, value=total)
        cell.number_format = MONEY_FMT
        cell.alignment = RIGHT_MONEY
    return not_mapped


def procesar_anexo(template_path, auxiliar_path, cfdi_path, mes, output_path):
    mes_col = MES_COL[mes.lower()]
    fondo_id = detect_fondo(auxiliar_path)
    if fondo_id in NOMBRE_FONDO:
        fondo_label, fondo_short = NOMBRE_FONDO[fondo_id]
    else:
        fondo_label = f'FONDO 8270-{fondo_id}'
        fondo_short = f'FONDO-{fondo_id}'
    rfc_map = build_rfc_map(cfdi_path) if cfdi_path else {}
    descs = extract_descriptions(auxiliar_path)
    records = extract_records(auxiliar_path, rfc_map)
    total = sum(r['cargo'] for r in records)
    sin_rfc_list = [r for r in records if not r['rfc'] and r['cap'] in (2000, 3000, 5000, 6000)]

    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    not_mapped = fill_egresos(wb['EGRESOS'], records, fondo_label, fondo_short, mes_col)
    by_cap = {}
    for r in records:
        by_cap.setdefault(r['cap'], []).append(r)
    for cap, (sheet_name, has_prov) in CAP_SHEET_MAP.items():
        if sheet_name in wb.sheetnames:
            if cap in by_cap:
                fill_detalle(wb[sheet_name], by_cap[cap], has_prov, descs, fondo_label, mes)
            else:
                clear_data_area(wb[sheet_name], start_row=17)
    wb.save(output_path)

    return {
        'fondo': fondo_label,
        'mes': mes,
        'total_movs': len(records),
        'total_importe': total,
        'sin_rfc': sin_rfc_list,
        'no_mapeadas': not_mapped,
    }


@app.route('/', methods=['GET'])
def index():
    return render_template('index.html', meses=list(MES_COL.keys()))


@app.route('/generar', methods=['POST'])
def generar():
    template_file = request.files.get('template')
    auxiliar_file = request.files.get('auxiliar')
    cfdi_file = request.files.get('cfdi')
    mes = request.form.get('mes', '').lower()

    if not template_file or not auxiliar_file or not mes:
        flash('Faltan archivos o el mes. Sube la plantilla, el auxiliar y selecciona el mes.', 'error')
        return redirect(url_for('index'))
    if mes not in MES_COL:
        flash(f'Mes inválido: {mes}', 'error')
        return redirect(url_for('index'))

    tmpdir = tempfile.mkdtemp(prefix='anexo5_')
    try:
        template_path = os.path.join(tmpdir, 'template.xlsx')
        auxiliar_path = os.path.join(tmpdir, 'auxiliar.xlsx')
        template_file.save(template_path)
        auxiliar_file.save(auxiliar_path)
        cfdi_path = None
        if cfdi_file and cfdi_file.filename:
            cfdi_path = os.path.join(tmpdir, 'cfdi.xlsx')
            cfdi_file.save(cfdi_path)
        output_path = os.path.join(tmpdir, f'Anexo5_{mes.upper()}_2025.xlsx')

        result = procesar_anexo(template_path, auxiliar_path, cfdi_path, mes, output_path)

        # Cargar bytes en memoria y devolver
        with open(output_path, 'rb') as f:
            data = io.BytesIO(f.read())
        data.seek(0)

        download_name = f"Anexo5_{mes.upper()}_2025_{result['fondo'].split()[1]}.xlsx"
        return send_file(
            data,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        flash(f'Error procesando los archivos: {e}', 'error')
        return redirect(url_for('index'))
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
